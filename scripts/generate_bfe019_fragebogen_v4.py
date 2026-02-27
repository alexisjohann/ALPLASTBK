#!/usr/bin/env python3
"""
BFE019 Fragebogen v4 Generator - Multi-Tab Excel
=================================================
Erstellt aus v3-Template (Modul 1) und altem Nullmessungs-Fragebogen
ein neues Excel mit einem Tab pro Modul (M1-M5).

Quellen:
- BFE019_Fragebogen_v3_Modul1_Gesamtmodernisierung.xlsx (Template)
- 260113_BFE_Nullmessung ECHfP_Fragebogen_Review intervista_sib.xlsx (Alt)

Transformation v3:
- S1: Stories neutralisiert (kein EnergieSchweiz im Intro)
- S4: Matrix-Format für Likert-Blöcke
- S5: Randomisierungshinweise
- S6: Quota-Variablen nach oben
- AI-04: T7 positive Formulierung
- AI-05: W4 "organisieren" Wording
- AI-06: T0 3-Stufen-Skala
- A1 → A1a/A1b Split
- I1 → I1a/I1b Split
- A5 → A5a/A5b Split
- BCM-Parameter Spalte
- Literatur Spalte
- Hinweise Spalte
"""

import os
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Pfade
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MESSKONZEPT_DIR = os.path.join(
    BASE_DIR,
    "data/dr-datareq/sources/clients/bfe/projects/2026_bfe_019_echfp/subprojects/messkonzept_2026"
)
V3_FILE = os.path.join(MESSKONZEPT_DIR, "BFE019_Fragebogen_v3_Modul1_Gesamtmodernisierung.xlsx")
OLD_FILE = os.path.join(MESSKONZEPT_DIR, "260113_BFE_Nullmessung ECHfP_Fragebogen_Review intervista_sib.xlsx")
OUTPUT_FILE = os.path.join(MESSKONZEPT_DIR, "BFE019_Fragebogen_v4_AllModule.xlsx")

# Styles
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='024079', end_color='024079', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='024079')
SECTION_FILL = PatternFill(start_color='E8EFF7', end_color='E8EFF7', fill_type='solid')  # v3-Template Farbe
SUBSECTION_FONT = Font(name='Calibri', bold=True, size=10, color='024079')  # Sub-Section: kein Fill
NORMAL_FONT = Font(name='Calibri', size=10)
LIT_FONT = Font(name='Calibri', size=9, color='549EDE')  # Literatur: hellblau, kleiner
HINT_FONT = Font(name='Calibri', size=9, color='666666', italic=True)  # Hinweise: grau, kursiv
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical='top')

# Spaltenbreiten
COL_WIDTHS = {
    'A': 14,   # Item-ID
    'B': 55,   # Fragetext
    'C': 45,   # Antwortoptionen
    'D': 16,   # Fragetyp
    'E': 30,   # BCM-Parameter
    'F': 35,   # Literatur
    'G': 35,   # Hinweise
}


# =============================================================================
# MODUL-DEFINITIONEN
# =============================================================================

# Modul-spezifische Wörter für Ersetzungen
MODULE_CONFIG = {
    2: {
        'name': 'Gebäudehülle',
        'tab_name': 'Modul 2 - Gebäudehülle',
        'topic_short': 'Erneuerung der Gebäudehülle',
        'topic_action': 'Erneuerung oder Dämmung der Gebäudehülle',
        'begriff_title': 'BEGRIFF_GH',
        'begriff_text': (
            'Mit «Erneuerung der Gebäudehülle» meinen wir alle Massnahmen, '
            'die die äussere Hülle Ihres Gebäudes betreffen: Fassadendämmung, '
            'Dachdämmung, Fensterersatz oder Kellerdeckendämmung. '
            'Ziel ist es, den Energieverbrauch zu senken und den Wohnkomfort zu steigern.'
        ),
        'story1_context': 'ob eine Erneuerung der Gebäudehülle sinnvoll wäre',
        'story2_context': 'Sie möchten die Gebäudehülle Ihres Hauses erneuern lassen',
        'story3_context': 'Informationen zur Erneuerung der Gebäudehülle',
        'w1_text': 'Ich bin bereit, in die Erneuerung der Gebäudehülle zu investieren.',
        'w2_text': 'Ich habe mich bereits über Möglichkeiten zur Erneuerung der Gebäudehülle informiert.',
        'w3_text': 'Ich habe bereits konkrete Schritte zur Erneuerung der Gebäudehülle unternommen.',
        'w4_text': 'Ich traue mir zu, eine Erneuerung der Gebäudehülle erfolgreich zu organisieren und in die Wege zu leiten.',
        'w5_text': 'Es fällt mir leicht, die Erneuerung oder Dämmung der Gebäudehülle zu planen und umzusetzen.',
        'w6_text': 'Ich habe das Gefühl, dass ich die Erneuerung der Gebäudehülle in nützlicher Frist umsetzen kann.',
        'i1a_text': 'Haben Sie schon einmal Förderprogramme für die Erneuerung der Gebäudehülle in Anspruch genommen oder sich darüber informiert?',
        'i1b_text': 'Haben Sie schon einmal Beratungsangebote für die Erneuerung der Gebäudehülle in Anspruch genommen oder sich darüber informiert?',
        'i2_text': 'Welche der folgenden Fachleute haben Sie in den letzten 12 Monaten zum Thema Gebäudehülle konsultiert?',
        'i2_options': 'Energieberater:in\nArchitekt:in\nGebäudehüllenspezialist:in\nDachdecker:in\nFensterbauer:in\nHandwerker:in\nKeiner / Keine\nAndere (bitte angeben)',
        't1_text': 'EnergieSchweiz bietet nützliche Informationen zur Erneuerung der Gebäudehülle.',
        't3_text': 'EnergieSchweiz hat dazu beigetragen, dass ich mich mit der Erneuerung der Gebäudehülle beschäftige.',
        't5_text': 'Es gibt genügend Informationen zu Förderprogrammen für die Erneuerung der Gebäudehülle.',
        't6_text': 'Ich wüsste, wo ich Unterstützung für die Erneuerung der Gebäudehülle finden kann.',
        'z2_items': (
            'Heiz- und Energiekosten senken\n'
            'Wohnkomfort und Behaglichkeit\n'
            'Langlebigkeit der Materialien\n'
            'Umwelt- und Klimaschutz\n'
            'Wertsteigerung der Immobilie\n'
            'Unabhängigkeit von Energiepreisen'
        ),
        'z3_text': 'Von welchen der folgenden Informationsquellen haben Sie in den letzten 12 Monaten Informationen zur Erneuerung der Gebäudehülle erhalten?',
        'z3_options': (
            'Internet / Websites\nSoziale Medien\nZeitungen / Zeitschriften\n'
            'TV / Radio\nFachgeschäft / Beratung\n'
            'Gebäudehüllenspezialist:in / Dachdecker:in / Fensterbauer:in\n'
            'Freunde / Familie / Nachbarn\n'
            'Gemeinde / Kanton\nEnergieberater:in\nAndere (bitte angeben)\nKeine'
        ),
        'z4_text': 'Was sind die grössten Hindernisse für die Erneuerung der Gebäudehülle bei Ihnen?',
        'z4_options': (
            'Zu hohe Kosten\nFehlende Fördermittel\n'
            'Bauliche Gegebenheiten (z.B. Denkmalschutz)\n'
            'Mangel an Fachleuten\nZu aufwändig / kompliziert\n'
            'Nicht genug informiert\nKein Bedarf\n'
            'Mietverhältnis (nicht Eigentümer:in)\n'
            'Andere (bitte angeben)'
        ),
        'has_campaign_questions': False,
    },
    3: {
        'name': 'Heizungsersatz',
        'tab_name': 'Modul 3 - Heizungsersatz',
        'topic_short': 'Erneuerung der Heizung',
        'topic_action': 'Heizungsersatz',
        'begriff_title': 'BEGRIFF_HZ',
        'begriff_text': (
            'Mit «Heizungsersatz» meinen wir den Wechsel von einem fossilen Heizsystem '
            '(Öl, Gas) zu einem erneuerbaren System wie Wärmepumpe, Holzheizung, '
            'Fernwärme oder Solarthermie. Ziel ist es, CO₂-Emissionen zu reduzieren '
            'und langfristig Heizkosten zu senken.'
        ),
        'story1_context': 'ob ein Heizungsersatz bei Ihnen sinnvoll wäre',
        'story2_context': 'Sie möchten Ihre Heizung durch ein erneuerbares System ersetzen',
        'story3_context': 'Informationen zum Heizungsersatz',
        'w1_text': 'Ich bin bereit, in die Erneuerung meiner Heizung zu investieren.',
        'w2_text': 'Ich habe mich bereits über Möglichkeiten zum Heizungsersatz informiert.',
        'w3_text': 'Ich habe bereits konkrete Schritte für einen Heizungsersatz unternommen.',
        'w4_text': 'Ich traue mir zu, die Erneuerung meiner Heizung erfolgreich zu organisieren und in die Wege zu leiten.',
        'w5_text': 'Es fällt mir leicht, einen Heizungsersatz zu planen und umzusetzen.',
        'w6_text': 'Ich habe das Gefühl, dass ich einen Heizungsersatz in nützlicher Frist umsetzen kann.',
        'i1a_text': 'Haben Sie schon einmal Förderprogramme für den Heizungsersatz in Anspruch genommen oder sich darüber informiert?',
        'i1b_text': 'Haben Sie schon einmal Beratungsangebote für den Heizungsersatz in Anspruch genommen oder sich darüber informiert?',
        'i2_text': 'Welche der folgenden Fachleute haben Sie in den letzten 12 Monaten zum Thema Heizungsersatz konsultiert?',
        'i2_options': 'Energieberater:in\nHeizungsinstallateur:in\nArchitekt:in\nHandwerker:in\nKeiner / Keine\nAndere (bitte angeben)',
        't1_text': 'EnergieSchweiz bietet nützliche Informationen zum Heizungsersatz.',
        't3_text': 'EnergieSchweiz hat dazu beigetragen, dass ich mich mit einem Heizungsersatz beschäftige.',
        't5_text': 'Es gibt genügend Informationen zu Förderprogrammen für den Heizungsersatz.',
        't6_text': 'Ich wüsste, wo ich Unterstützung für einen Heizungsersatz finden kann.',
        'z2_items': (
            'Heizkosten senken\n'
            'Heizleistung und Zuverlässigkeit\n'
            'Versorgungssicherheit\n'
            'Umwelt- und Klimaschutz\n'
            'Wertsteigerung der Immobilie\n'
            'Unabhängigkeit von fossilen Energien'
        ),
        'z3_text': 'Von welchen der folgenden Informationsquellen haben Sie in den letzten 12 Monaten Informationen zum Heizungsersatz erhalten?',
        'z3_options': (
            'Internet / Websites\nSoziale Medien\nZeitungen / Zeitschriften\n'
            'TV / Radio\nFachgeschäft / Beratung\n'
            'Heizungsinstallateur:in\n'
            'Freunde / Familie / Nachbarn\n'
            'Gemeinde / Kanton\nEnergieberater:in\nAndere (bitte angeben)\nKeine'
        ),
        'z4_text': 'Was sind die grössten Hindernisse für einen Heizungsersatz bei Ihnen?',
        'z4_options': (
            'Zu hohe Kosten\nFehlende Fördermittel\n'
            'Aktuelle Heizung funktioniert noch\n'
            'Mangel an Fachleuten\nZu aufwändig / kompliziert\n'
            'Nicht genug informiert\nKein Bedarf\n'
            'Mietverhältnis (nicht Eigentümer:in)\n'
            'Regionaler Heizstoff (z.B. günstiges Holz)\n'
            'Andere (bitte angeben)'
        ),
        'has_campaign_questions': True,
        'campaign_questions': [
            {
                'id': 'Z5',
                'text': 'Kennen Sie die Kampagne «erneuerbar heizen»?',
                'options': 'Ja, kenne ich\nSchon davon gehört, aber weiss nicht genau\nNein, noch nie davon gehört',
                'type': 'Einfachauswahl',
                'param': 'φ (BCJ-Phase): Awareness Kampagne',
                'lit': '',
                'hint': 'Kampagnen-spezifisch: «erneuerbar heizen»',
            },
            {
                'id': 'Z6',
                'text': 'Haben Sie die Impulsberatung «erneuerbar heizen» schon einmal in Anspruch genommen?',
                'options': 'Ja\nNein, aber ich kenne das Angebot\nNein, kenne ich nicht',
                'type': 'Einfachauswahl',
                'param': 'θ (Activation): Impulsberatung',
                'lit': '',
                'hint': 'Kampagnen-spezifisch',
            },
            {
                'id': 'Z7',
                'text': 'Wie wahrscheinlich ist es, dass Sie die Impulsberatung «erneuerbar heizen» in den nächsten 12 Monaten in Anspruch nehmen?',
                'options': 'Sehr wahrscheinlich [5]\nEher wahrscheinlich [4]\nTeils/teils [3]\nEher unwahrscheinlich [2]\nSehr unwahrscheinlich [1]',
                'type': 'Likert 5',
                'param': 'β (Present Bias): Intention → Verhalten',
                'lit': '',
                'hint': 'Intention-Behavior Gap messen',
            },
            {
                'id': 'Z8',
                'text': 'Falls Sie die Impulsberatung nicht in Anspruch genommen haben: Was hält Sie davon ab?',
                'options': 'Kein Bedarf\nZu teuer\nZu aufwändig\nKenne das Angebot nicht genug\nHabe schon andere Beratung\nAndere (bitte angeben)',
                'type': 'Multiple Choice',
                'param': 'τ (Complexity): Wahrgenommene Barrieren',
                'lit': '',
                'hint': 'Barrier-Analyse',
            },
            {
                'id': 'Z9',
                'text': 'Was müsste sich ändern, damit Sie die Impulsberatung nutzen würden?',
                'options': 'Offene Frage',
                'type': 'Offen',
                'param': '',
                'lit': '',
                'hint': 'Qualitativ: Verbesserungspotenzial',
            },
        ],
    },
    4: {
        'name': 'Solarenergie',
        'tab_name': 'Modul 4 - Solarenergie',
        'topic_short': 'Installation einer Solaranlage',
        'topic_action': 'Solarenergie / Photovoltaik',
        'begriff_title': 'BEGRIFF_PV',
        'begriff_text': (
            'Mit «Solarenergie» meinen wir die Nutzung von Sonnenenergie zur '
            'Stromerzeugung (Photovoltaik) oder Wärmegewinnung (Solarthermie) '
            'auf Ihrem Gebäude. Ziel ist es, eigenen Strom zu produzieren, '
            'Energiekosten zu senken und einen Beitrag zum Klimaschutz zu leisten.'
        ),
        'story1_context': 'ob eine Solaranlage auf Ihrem Dach sinnvoll wäre',
        'story2_context': 'Sie möchten eine Solaranlage auf Ihrem Dach installieren lassen',
        'story3_context': 'Informationen zur Installation einer Solaranlage',
        'w1_text': 'Ich bin bereit, in die Installation einer Solaranlage zu investieren.',
        'w2_text': 'Ich habe mich bereits über Möglichkeiten zur Solarenergienutzung informiert.',
        'w3_text': 'Ich habe bereits konkrete Schritte zur Installation einer Solaranlage unternommen.',
        'w4_text': 'Ich traue mir zu, die Installation einer Solaranlage erfolgreich zu organisieren und in die Wege zu leiten.',
        'w5_text': 'Es fällt mir leicht, eine Solaranlage zur Stromerzeugung zu planen und installieren zu lassen.',
        'w6_text': 'Ich habe das Gefühl, dass ich die Installation einer Solaranlage in nützlicher Frist umsetzen kann.',
        'i1a_text': 'Haben Sie schon einmal Förderprogramme für Solaranlagen in Anspruch genommen oder sich darüber informiert?',
        'i1b_text': 'Haben Sie schon einmal Beratungsangebote für Solaranlagen in Anspruch genommen oder sich darüber informiert?',
        'i2_text': 'Welche der folgenden Fachleute haben Sie in den letzten 12 Monaten zum Thema Solarenergie konsultiert?',
        'i2_options': 'Energieberater:in\nSolarinstallateur:in\nElektriker:in\nArchitekt:in\nKeiner / Keine\nAndere (bitte angeben)',
        't1_text': 'EnergieSchweiz bietet nützliche Informationen zur Solarenergie.',
        't3_text': 'EnergieSchweiz hat dazu beigetragen, dass ich mich mit Solarenergie beschäftige.',
        't5_text': 'Es gibt genügend Informationen zu Förderprogrammen für Solaranlagen.',
        't6_text': 'Ich wüsste, wo ich Unterstützung für die Installation einer Solaranlage finden kann.',
        'z2_items': (
            'Eigenverbrauch des Solarstroms\n'
            'Schnelle finanzielle Rentabilität\n'
            'Lebensdauer und Wartungsarmut\n'
            'Umwelt- und Klimaschutz\n'
            'Wertsteigerung der Immobilie\n'
            'Unabhängigkeit vom Stromnetz'
        ),
        'z3_text': 'Von welchen der folgenden Informationsquellen haben Sie in den letzten 12 Monaten Informationen zur Solarenergie erhalten?',
        'z3_options': (
            'Internet / Websites\nSoziale Medien\nZeitungen / Zeitschriften\n'
            'TV / Radio\nFachgeschäft / Beratung\n'
            'Solarinstallateur:in\n'
            'Freunde / Familie / Nachbarn\n'
            'Gemeinde / Kanton\nEnergieberater:in\nSolarrechner (online)\nAndere (bitte angeben)\nKeine'
        ),
        'z4_text': 'Was sind die grössten Hindernisse für die Installation einer Solaranlage bei Ihnen?',
        'z4_options': (
            'Zu hohe Kosten\nFehlende Fördermittel\n'
            'Dach nicht geeignet\nBauliche Gegebenheiten (z.B. Denkmalschutz)\n'
            'Mangel an Fachleuten\nZu aufwändig / kompliziert\n'
            'Nicht genug informiert\nKein Bedarf\n'
            'Mietverhältnis (nicht Eigentümer:in)\n'
            'Andere (bitte angeben)'
        ),
        'has_campaign_questions': True,
        'campaign_questions': [
            {
                'id': 'Z5',
                'text': 'Wissen Sie, wie viel Solarpotenzial Ihr Dach hat bzw. wie viel Strom Sie damit erzeugen könnten?',
                'options': 'Ja, ich kenne das Potenzial\nUngefähr\nNein',
                'type': 'Einfachauswahl',
                'param': 'A(·): Awareness Potenzial',
                'lit': '',
                'hint': 'Solar-spezifisch: Potenzialkenntnis',
            },
            {
                'id': 'Z6',
                'text': 'Haben Sie bereits eine Batterie / einen Stromspeicher oder planen Sie die Anschaffung?',
                'options': 'Ja, habe bereits\nJa, plane Anschaffung\nNein, kein Interesse\nKenne mich nicht aus',
                'type': 'Einfachauswahl',
                'param': 'θ (Activation): Batterie-Adoption',
                'lit': '',
                'hint': 'Solar-spezifisch',
            },
            {
                'id': 'Z7',
                'text': 'Sind Sie Mitglied in einem Zusammenschluss zum Eigenverbrauch (ZEV)?',
                'options': 'Ja\nNein, aber interessiert\nNein, kein Interesse\nKenne ich nicht',
                'type': 'Einfachauswahl',
                'param': 'σ (Social): Kooperatives Verhalten',
                'lit': '',
                'hint': 'Solar-spezifisch: ZEV',
            },
            {
                'id': 'Z8',
                'text': 'Kennen Sie Fördermöglichkeiten für Solaranlagen in Ihrem Kanton?',
                'options': 'Ja, kenne ich gut\nUngefähr\nNein',
                'type': 'Einfachauswahl',
                'param': 'A(·): Awareness Förderung',
                'lit': '',
                'hint': 'Solar-spezifisch',
            },
            {
                'id': 'Z9',
                'text': 'Haben Sie schon einmal einen Solarrechner (z.B. sonnendach.ch) genutzt?',
                'options': 'Ja\nNein, aber ich kenne das Angebot\nNein, kenne ich nicht',
                'type': 'Einfachauswahl',
                'param': 'θ (Activation): Tool-Nutzung',
                'lit': '',
                'hint': 'Solar-spezifisch: Solarrechner',
            },
        ],
    },
    5: {
        'name': 'E-Mobilität',
        'tab_name': 'Modul 5 - E-Mobilität',
        'topic_short': 'Installation der Ladeinfrastruktur',
        'topic_action': 'Elektromobilität / Ladeinfrastruktur',
        'begriff_title': 'BEGRIFF_EM',
        'begriff_text': (
            'Mit «E-Mobilität» meinen wir die Nutzung von Elektrofahrzeugen '
            'und die Installation der zugehörigen Ladeinfrastruktur (z.B. Wallbox) '
            'an Ihrem Wohnort. Ziel ist es, die Mobilität klimafreundlicher '
            'zu gestalten und von tieferen Betriebskosten zu profitieren.'
        ),
        'story1_context': 'ob E-Mobilität und eine Ladestation bei Ihnen sinnvoll wären',
        'story2_context': 'Sie möchten eine Ladestation für Elektrofahrzeuge installieren lassen',
        'story3_context': 'Informationen zur Elektromobilität und Ladeinfrastruktur',
        'w1_text': 'Ich bin bereit, in E-Mobilität und Ladeinfrastruktur zu investieren.',
        'w2_text': 'Ich habe mich bereits über Möglichkeiten zur Elektromobilität informiert.',
        'w3_text': 'Ich habe bereits konkrete Schritte in Richtung E-Mobilität unternommen.',
        'w4_text': 'Ich traue mir zu, die Installation der Ladeinfrastruktur erfolgreich zu organisieren und in die Wege zu leiten.',
        'w5_text': 'Es fällt mir leicht, eine Ladeinfrastruktur für Elektrofahrzeuge zu planen und installieren zu lassen.',
        'w6_text': 'Ich habe das Gefühl, dass ich die Installation einer Ladestation in nützlicher Frist umsetzen kann.',
        'i1a_text': 'Haben Sie schon einmal Förderprogramme für Elektromobilität in Anspruch genommen oder sich darüber informiert?',
        'i1b_text': 'Haben Sie schon einmal Beratungsangebote für Elektromobilität in Anspruch genommen oder sich darüber informiert?',
        'i2_text': 'Welche der folgenden Fachleute haben Sie in den letzten 12 Monaten zum Thema E-Mobilität konsultiert?',
        'i2_options': 'Energieberater:in\nElektriker:in\nAuto-Händler:in\nArchitekt:in\nKeiner / Keine\nAndere (bitte angeben)',
        't1_text': 'EnergieSchweiz bietet nützliche Informationen zur E-Mobilität.',
        't3_text': 'EnergieSchweiz hat dazu beigetragen, dass ich mich mit E-Mobilität beschäftige.',
        't5_text': 'Es gibt genügend Informationen zu Förderprogrammen für E-Mobilität.',
        't6_text': 'Ich wüsste, wo ich Unterstützung für die Installation einer Ladestation finden kann.',
        'z2_items': (
            'Betriebskosten senken\n'
            'Zukunftssicherheit der Technologie\n'
            'Laden mit eigener PV-Anlage\n'
            'Umwelt- und Klimaschutz\n'
            'Fahrkomfort und Leistung\n'
            'Unabhängigkeit von fossilen Treibstoffen'
        ),
        'z3_text': 'Von welchen der folgenden Informationsquellen haben Sie in den letzten 12 Monaten Informationen zur E-Mobilität erhalten?',
        'z3_options': (
            'Internet / Websites\nSoziale Medien\nZeitungen / Zeitschriften\n'
            'TV / Radio\nAutohaus / Händler\n'
            'Elektriker:in\n'
            'Freunde / Familie / Nachbarn\n'
            'Gemeinde / Kanton\nEnergieberater:in\nAndere (bitte angeben)\nKeine'
        ),
        'z4_text': 'Was sind die grössten Hindernisse für E-Mobilität und Ladeinfrastruktur bei Ihnen?',
        'z4_options': (
            'Zu hohe Kosten (Fahrzeug)\nZu hohe Kosten (Ladestation)\n'
            'Fehlende Lademöglichkeit am Wohnort\n'
            'Reichweite der Fahrzeuge\nMangel an öffentlichen Ladestationen\n'
            'Zu aufwändig / kompliziert\n'
            'Nicht genug informiert\nKein Bedarf\n'
            'Mietverhältnis (nicht Eigentümer:in)\n'
            'Andere (bitte angeben)'
        ),
        'has_campaign_questions': True,
        'campaign_questions': [
            {
                'id': 'Z5',
                'text': 'Wissen Sie, ob an Ihrem Wohnort die Installation einer Wallbox möglich wäre?',
                'options': 'Ja, ist möglich\nUnsicher\nNein, nicht möglich\nWeiss nicht',
                'type': 'Einfachauswahl',
                'param': 'A(·): Awareness Möglichkeit',
                'lit': '',
                'hint': 'E-Mob-spezifisch: Wallbox-Möglichkeit',
            },
            {
                'id': 'Z6',
                'text': 'Wissen Sie, wie Sie bei der Installation einer Ladestation vorgehen müssten?',
                'options': 'Ja, bin gut informiert\nUngefähr\nNein',
                'type': 'Einfachauswahl',
                'param': 'τ (Complexity): Prozesswissen',
                'lit': '',
                'hint': 'E-Mob-spezifisch',
            },
            {
                'id': 'Z7',
                'text': 'Kennen Sie die Kampagne «Fahr mit dem Strom»?',
                'options': 'Ja, kenne ich\nSchon davon gehört, aber weiss nicht genau\nNein, noch nie davon gehört',
                'type': 'Einfachauswahl',
                'param': 'φ (BCJ-Phase): Awareness Kampagne',
                'lit': '',
                'hint': 'Kampagnen-spezifisch: «Fahr mit dem Strom»',
            },
        ],
    },
}

# BCM-Parameter Mapping (identisch für alle Module, A-W-I-T Struktur)
BCM_PARAMS = {
    'A1a': 'A(·): Gestützte Awareness (Informationen)',
    'A1b': 'A(·): Gestützte Awareness (Werbung)',
    'A2': 'A(·): Thematische Salienz (Beschäftigung)',
    'A3': 'A(·): Persönliche Betroffenheit',
    'A4': 'A(·): Subjektives Wissen / Selbsteinschätzung',
    'A5a': 'σ (Social Norms): Wahrgenommene gesellschaftliche Bedeutung',
    'A5b': 'u_E (Environmental Utility): Umweltbedeutung des eigenen Verhaltens',
    'W1': 'β (Present Bias): Investitionsbereitschaft',
    'W2': 'θ (Activation): Informationsverhalten',
    'W3': 'θ (Activation): Konkrete Handlungsschritte',
    'W4': 'τ (Complexity): Self-Efficacy (Selbstwirksamkeit)',
    'W5': 'τ (Complexity): Perceived Ease (Wahrgenommene Einfachheit)',
    'W6': 'β (Present Bias): Zeitliche Umsetzbarkeit',
    'I1a': 'θ (Activation): Inanspruchnahme Förderprogramme',
    'I1b': 'θ (Activation): Inanspruchnahme Beratung',
    'I2': 'θ (Activation): Fachleute-Konsultation',
    'T0': 'A(·): Gestützte Awareness EnergieSchweiz',
    'T1': 'Ψ (Trust): Wahrgenommener Nutzen ES',
    'T2': 'Ψ (Trust): Vertrauenswürdigkeit ES',
    'T3': 'Ψ (Trust): Einfluss ES auf eigenes Verhalten',
    'T4': 'Ψ (Trust): Kompetenz ES',
    'T5': 'A(·): Wahrgenommene Informationsverfügbarkeit Förderung',
    'T6': 'θ (Activation): Wissen über Unterstützungsangebote',
    'T7': 'Ψ (Trust): Verlässlichkeit ES-Informationen',
    'Z1': 'u_FEPSDE: Relative Wichtigkeit Utility-Dimensionen',
    'Z2': 'u_FEPSDE: MaxDiff Utility-Gewichte',
    'Z3': 'A(·): Informationsquellen-Nutzung',
    'Z4': 'τ (Complexity): Wahrgenommene Barrieren',
}

# Literatur Mapping (identisch für alle Module)
LITERATUR = {
    'A1a': 'Keller & Lehmann (2006); Alba & Hutchinson (1987)',
    'A1b': 'Keller & Lehmann (2006); Alba & Hutchinson (1987)',
    'A2': 'Kahneman (2011) Salience; Bordalo et al. (2013)',
    'A3': 'Slovic (2000) Affect Heuristic; Loewenstein et al. (2001)',
    'A4': 'Kruger & Dunning (1999); Alba & Hutchinson (2000)',
    'A5a': 'Cialdini et al. (1990) Social Norms; Schultz et al. (2007)',
    'A5b': 'Stern (2000) VBN Theory; Steg & Vlek (2009)',
    'W1': 'Thaler (1981) Temporal Discounting; Frederick et al. (2002)',
    'W2': 'Rogers (2003) Diffusion of Innovations',
    'W3': 'Gollwitzer (1999) Implementation Intentions',
    'W4': 'Bandura (1977) Self-Efficacy; Ajzen (1991) PBC',
    'W5': 'Davis (1989) TAM; Venkatesh et al. (2003) UTAUT',
    'W6': 'Ariely & Wertenbroch (2002) Procrastination',
    'I1a': 'Allcott & Greenstone (2012); Fowlie et al. (2018)',
    'I1b': 'Allcott & Greenstone (2012); Fowlie et al. (2018)',
    'I2': 'Joskow & Marron (1992); Stern et al. (2016)',
    'T0': 'Keller & Lehmann (2006) Brand Awareness',
    'T1': 'Mayer et al. (1995) Trust Model',
    'T2': 'Mayer et al. (1995) Trust Model',
    'T3': 'Cialdini (2006) Authority Principle',
    'T4': 'Mayer et al. (1995) Ability Component',
    'T5': 'Thaler & Sunstein (2008) Choice Architecture',
    'T6': 'Rogers (2003) Innovation-Decision Process',
    'T7': 'Mayer et al. (1995) Integrity Component',
    'Z1': 'Lancaster (1966) Characteristics Theory; Fehr EBF FEPSDE',
    'Z2': 'Louviere et al. (2015) MaxDiff; Sawtooth Software',
    'Z3': 'Rogers (2003) Communication Channels',
    'Z4': 'Stern (2000); Steg et al. (2005) Barrier Model',
}


def apply_styles(ws, row, is_header=False, is_section=False, is_subsection=False):
    """Wende Styles auf eine Zeile an."""
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGNMENT
        if is_header:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        elif is_section:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        elif is_subsection:
            cell.font = SUBSECTION_FONT
        else:
            # Spalten-spezifische Fonts für F (Literatur) und G (Hinweise)
            if col == 6:  # Literatur
                cell.font = LIT_FONT
            elif col == 7:  # Hinweise
                cell.font = HINT_FONT
            else:
                cell.font = NORMAL_FONT


def set_col_widths(ws):
    """Setze Spaltenbreiten."""
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def write_header(ws, row=1):
    """Schreibe Header-Zeile."""
    headers = ['Item-ID', 'Fragetext', 'Antwortoptionen', 'Fragetyp',
               'BCM-Parameter', 'Literatur', 'Hinweise']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    apply_styles(ws, row, is_header=True)
    return row + 1


def write_section(ws, row, title):
    """Schreibe Abschnitts-Header."""
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    apply_styles(ws, row, is_section=True)
    return row + 1


def write_subsection(ws, row, item_id, text, ftype='', hint=''):
    """Schreibe eine Sub-Section-Zeile (STORY, MATRIX_INTRO, BEGRIFF) — v3-Stil: kein Fill, merged."""
    ws.cell(row=row, column=1, value=item_id)
    ws.cell(row=row, column=2, value=text)
    ws.cell(row=row, column=4, value=ftype)
    ws.cell(row=row, column=7, value=hint)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    apply_styles(ws, row, is_subsection=True)
    # Hinweise-Spalte bleibt kursiv/grau auch bei Sub-Sections
    ws.cell(row=row, column=7).font = HINT_FONT
    return row + 1


def write_row(ws, row, item_id='', text='', options='', ftype='', param='', lit='', hint=''):
    """Schreibe eine Datenzeile."""
    ws.cell(row=row, column=1, value=item_id)
    ws.cell(row=row, column=2, value=text)
    ws.cell(row=row, column=3, value=options)
    ws.cell(row=row, column=4, value=ftype)
    ws.cell(row=row, column=5, value=param)
    ws.cell(row=row, column=6, value=lit)
    ws.cell(row=row, column=7, value=hint)
    apply_styles(ws, row)
    return row + 1


def write_likert5_options():
    """Standard Likert-5 Antwortoptionen."""
    return 'Trifft voll zu [5]\nTrifft eher zu [4]\nTeils/teils [3]\nTrifft eher nicht zu [2]\nTrifft gar nicht zu [1]'


def copy_v3_template(wb_new, v3_wb):
    """Kopiere Modul 1 aus v3-Template komplett."""
    v3_ws = v3_wb['Modul 1 - Gesamtmodernisierung']
    ws = wb_new.active
    ws.title = 'Modul 1 - Gesamtmodernisierung'

    for row in range(1, v3_ws.max_row + 1):
        for col in range(1, v3_ws.max_column + 1):
            src_cell = v3_ws.cell(row=row, column=col)
            dst_cell = ws.cell(row=row, column=col, value=src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.number_format = src_cell.number_format

    # Kopiere Spaltenbreiten
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        if col_letter in v3_ws.column_dimensions:
            ws.column_dimensions[col_letter].width = v3_ws.column_dimensions[col_letter].width

    # Kopiere Merged Cells
    for merge in v3_ws.merged_cells.ranges:
        ws.merge_cells(str(merge))

    return ws


def create_module_tab(wb, mod_num, config):
    """Erstelle einen Modul-Tab mit v3-Struktur."""
    ws = wb.create_sheet(title=config['tab_name'])
    set_col_widths(ws)
    likert5 = write_likert5_options()

    r = write_header(ws)

    # -- WILLKOMMEN --
    r = write_section(ws, r, f'WILLKOMMEN — {config["name"]}')
    r = write_subsection(ws, r,
        item_id='WILLKOMMEN',
        text=(
            f'Herzlich willkommen! In dieser Befragung geht es um das Thema '
            f'«{config["topic_action"]}». Ihre Meinung ist uns wichtig. '
            f'Die Befragung dauert ca. 10 Minuten. Alle Angaben sind anonym.'
        ),
        ftype='Text',
        hint='Neutralisiert (S1): Kein EnergieSchweiz im Intro'
    )

    # -- STORY 1 (neutralisiert) --
    r = write_section(ws, r, f'AWARENESS — {config["name"]}')
    r = write_subsection(ws, r,
        item_id='STORY_1',
        text=(
            f'Stellen Sie sich vor: Sie denken über Ihr Haus oder Ihre Wohnung nach und '
            f'{config["story1_context"]} für die Zukunft. '
            f'Was wissen Sie darüber, welche Erfahrungen haben Sie gemacht?'
        ),
        ftype='Text (Überleitung)',
        hint='Neutralisiert (S1): Keine Marke, kein Storytelling-Risiko'
    )

    # -- BEGRIFF --
    r = write_subsection(ws, r,
        item_id=config['begriff_title'],
        text=config['begriff_text'],
        ftype='Definition (eingeblendet)',
        hint='Begriffserklärung für Befragte'
    )

    # -- AWARENESS Block (A1a-A5b) --
    r = write_subsection(ws, r,
        item_id='[MATRIX_INTRO_A]',
        text='Inwiefern treffen die folgenden Aussagen auf Sie zu?',
        ftype='Matrix-Intro',
        hint='S4: Matrix-Format für Likert-Block'
    )
    r = write_row(ws, r, 'A1a',
        f'Ich habe in den letzten 12 Monaten Informationen zum Thema «{config["topic_short"]}» wahrgenommen.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A1a'], LITERATUR['A1a'],
        'Split aus A1 (v3): Informationen separat')
    r = write_row(ws, r, 'A1b',
        f'Ich habe in den letzten 12 Monaten Werbung zum Thema «{config["topic_short"]}» wahrgenommen.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A1b'], LITERATUR['A1b'],
        'Split aus A1 (v3): Werbung separat')
    r = write_row(ws, r, 'A2',
        f'Das Thema «{config["topic_short"]}» beschäftigt mich.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A2'], LITERATUR['A2'], '')
    r = write_row(ws, r, 'A3',
        f'Das Thema «{config["topic_short"]}» betrifft mich persönlich.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A3'], LITERATUR['A3'], '')
    r = write_row(ws, r, 'A4',
        f'Beim Thema «{config["topic_short"]}» weiss ich, um was es geht.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A4'], LITERATUR['A4'], '')
    r = write_row(ws, r, 'A5a',
        'Mein eigenes Verhalten ist von Bedeutung für die Gesellschaft.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A5a'], LITERATUR['A5a'],
        'Split aus A5 (v3): Gesellschaft separat')
    r = write_row(ws, r, 'A5b',
        'Mein eigenes Verhalten ist von Bedeutung für die Umwelt.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['A5b'], LITERATUR['A5b'],
        'Split aus A5 (v3): Umwelt separat')

    # -- WILLINGNESS Block (W1-W6) --
    r = write_section(ws, r, f'WILLINGNESS — {config["name"]}')
    r = write_subsection(ws, r,
        item_id='[MATRIX_INTRO_W]',
        text='Inwiefern treffen die folgenden Aussagen auf Sie zu?',
        ftype='Matrix-Intro',
        hint='S4: Matrix-Format für Likert-Block'
    )
    r = write_row(ws, r, 'W1', config['w1_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['W1'], LITERATUR['W1'], '')
    r = write_row(ws, r, 'W2', config['w2_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['W2'], LITERATUR['W2'], '')
    r = write_row(ws, r, 'W3', config['w3_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['W3'], LITERATUR['W3'], '')
    r = write_row(ws, r, 'W4', config['w4_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['W4'], LITERATUR['W4'], 'AI-05: "organisieren und in die Wege zu leiten"')
    r = write_row(ws, r, 'W5', config['w5_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['W5'], LITERATUR['W5'], '')
    r = write_row(ws, r, 'W6', config['w6_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['W6'], LITERATUR['W6'], '')

    # -- IMPACT Block --
    r = write_section(ws, r, f'IMPACT — {config["name"]}')
    r = write_subsection(ws, r,
        item_id='STORY_2',
        text=(
            f'Denken Sie nun an Ihr Haus oder Ihre Wohnung. Stellen Sie sich vor, '
            f'{config["story2_context"]}.'
        ),
        ftype='Text (Überleitung)',
        hint='Neutralisiert (S1)'
    )
    r = write_row(ws, r, 'I1a', config['i1a_text'],
        'Ja, in Anspruch genommen\nJa, mich informiert\nNein',
        'Einfachauswahl', BCM_PARAMS['I1a'], LITERATUR['I1a'],
        'Split aus I1 (v3): Förderprogramme separat')
    r = write_row(ws, r, 'I1b', config['i1b_text'],
        'Ja, in Anspruch genommen\nJa, mich informiert\nNein',
        'Einfachauswahl', BCM_PARAMS['I1b'], LITERATUR['I1b'],
        'Split aus I1 (v3): Beratung separat')
    r = write_row(ws, r, 'I2', config['i2_text'],
        config['i2_options'],
        'Multiple Choice', BCM_PARAMS['I2'], LITERATUR['I2'], '')

    # -- TRUST Block --
    r = write_section(ws, r, f'TRUST (EnergieSchweiz) — {config["name"]}')
    r = write_subsection(ws, r,
        item_id='STORY_3',
        text=(
            f'Stellen Sie sich nun vor, Sie suchen im Internet nach '
            f'{config["story3_context"]}. '
            f'Sie stossen dabei auf die Website von EnergieSchweiz.'
        ),
        ftype='Text (Überleitung)',
        hint='Neutralisiert (S1)'
    )
    r = write_subsection(ws, r,
        item_id='BEGRIFF_ES',
        text=(
            'EnergieSchweiz ist das Programm des Bundesamtes für Energie (BFE) '
            'für Energieeffizienz und erneuerbare Energien. Es informiert, '
            'berät und sensibilisiert die Bevölkerung.'
        ),
        ftype='Definition (eingeblendet)',
        hint='Begriffserklärung EnergieSchweiz'
    )
    r = write_row(ws, r, 'T0',
        'Kennen Sie EnergieSchweiz?',
        'Noch nie davon gehört [1]\nSchon davon gehört, aber weiss nicht genau [2]\nKenne EnergieSchweiz [3]',
        'Einfachauswahl (3-stufig)', BCM_PARAMS['T0'], LITERATUR['T0'],
        'AI-06: 3-Stufen-Skala statt Ja/Nein')

    # T1-T7 als Matrix
    r = write_subsection(ws, r,
        item_id='[MATRIX_INTRO_T]',
        text='Inwiefern treffen die folgenden Aussagen auf EnergieSchweiz zu?',
        ftype='Matrix-Intro',
        hint='S4: Matrix-Format; Filter: Nur wenn T0 ≥ 2'
    )
    r = write_row(ws, r, 'T1', config['t1_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['T1'], LITERATUR['T1'], '')
    r = write_row(ws, r, 'T2',
        'Ich vertraue den Informationen von EnergieSchweiz.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['T2'], LITERATUR['T2'], '')
    r = write_row(ws, r, 'T3', config['t3_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['T3'], LITERATUR['T3'], '')
    r = write_row(ws, r, 'T4',
        'EnergieSchweiz ist kompetent im Bereich Energie.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['T4'], LITERATUR['T4'], '')
    r = write_row(ws, r, 'T5', config['t5_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['T5'], LITERATUR['T5'], '')
    r = write_row(ws, r, 'T6', config['t6_text'], likert5, 'Likert (Matrix)',
        BCM_PARAMS['T6'], LITERATUR['T6'], '')
    r = write_row(ws, r, 'T7',
        'Ich kann mich auf die Informationen von EnergieSchweiz verlassen.',
        likert5, 'Likert (Matrix)', BCM_PARAMS['T7'], LITERATUR['T7'],
        'AI-04: Positive Formulierung (statt invertiert)')

    # -- ZUSATZFRAGEN --
    r = write_section(ws, r, f'ZUSATZFRAGEN — {config["name"]}')

    # Z1: Likert Matrix (Wichtigkeit Faktoren)
    r = write_subsection(ws, r,
        item_id='[MATRIX_INTRO_Z1]',
        text=f'Wie wichtig sind Ihnen die folgenden Aspekte bei der Entscheidung für eine «{config["topic_short"]}»?',
        ftype='Matrix-Intro',
        hint='S4: Matrix-Format; geändert von Rangfolge zu Likert'
    )
    z1_options = 'Sehr wichtig [5]\nEher wichtig [4]\nTeils/teils [3]\nEher unwichtig [2]\nGar nicht wichtig [1]'
    r = write_row(ws, r, 'Z1', 'Kosten und Finanzierung', z1_options,
        'Likert (Matrix)', BCM_PARAMS['Z1'], LITERATUR['Z1'],
        'FEPSDE: u_F')
    r = write_row(ws, r, '', 'Umwelt- und Klimaschutz', z1_options,
        'Likert (Matrix)', '', '', 'FEPSDE: u_E')
    r = write_row(ws, r, '', 'Komfort und Lebensqualität', z1_options,
        'Likert (Matrix)', '', '', 'FEPSDE: u_P')
    r = write_row(ws, r, '', 'Wertsteigerung der Immobilie', z1_options,
        'Likert (Matrix)', '', '', 'FEPSDE: u_F')
    r = write_row(ws, r, '', 'Was Nachbarn / Bekannte tun', z1_options,
        'Likert (Matrix)', '', '', 'FEPSDE: u_S')
    r = write_row(ws, r, '', 'Unabhängigkeit', z1_options,
        'Likert (Matrix)', '', '', 'FEPSDE: u_X')

    # Z2: MaxDiff
    r = write_row(ws, r, 'Z2',
        f'Stellen Sie sich vor, Sie müssen sich für eine «{config["topic_short"]}» entscheiden. '
        f'Welcher der folgenden Aspekte wäre Ihnen am WICHTIGSTEN und welcher am UNWICHTIGSTEN?',
        config['z2_items'],
        'MaxDiff (Best-Worst)', BCM_PARAMS['Z2'], LITERATUR['Z2'],
        'MaxDiff-Design: Utility-Gewichte u_FEPSDE')

    # Z3: Multiple Choice (Informationsquellen)
    r = write_row(ws, r, 'Z3', config['z3_text'],
        config['z3_options'],
        'Multiple Choice', BCM_PARAMS['Z3'], LITERATUR['Z3'],
        'S5: Items randomisiert (ausser «Keine»)')

    # Z4: Multiple Choice (Hindernisse)
    r = write_row(ws, r, 'Z4', config['z4_text'],
        config['z4_options'],
        'Multiple Choice', BCM_PARAMS['Z4'], LITERATUR['Z4'],
        'S5: Items randomisiert (ausser «Andere»)')

    # Kampagnen-spezifische Zusatzfragen (Z5-Z9)
    if config.get('has_campaign_questions') and config.get('campaign_questions'):
        r = write_section(ws, r, f'KAMPAGNEN-SPEZIFISCH — {config["name"]}')
        for cq in config['campaign_questions']:
            r = write_row(ws, r, cq['id'], cq['text'], cq['options'],
                cq['type'], cq.get('param', ''), cq.get('lit', ''), cq.get('hint', ''))

    return ws


def create_aenderungsprotokoll(wb):
    """Erstelle das Änderungsprotokoll-Tab."""
    ws = wb.create_sheet(title='Änderungsprotokoll')
    headers = ['Änderung', 'Item(s)', 'Beschreibung', 'Quelle']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 25

    changes = [
        ['S1', 'STORY_1/2/3', 'Stories neutralisiert: Kein EnergieSchweiz/Markennennung im Intro und Überleitungen. Baseline-Messung darf keine Intervention sein.', 'Meeting 2026-02-18 (Abi)'],
        ['S4', 'A-Block, W-Block, T-Block', 'Likert-Fragen im Matrix-Format gruppiert mit Matrix-Intro', 'Best Practice Fragebogendesign'],
        ['S5', 'Z3, Z4', 'Randomisierung der Items (ausser Ankeritems wie «Keine»/«Andere»)', 'Best Practice'],
        ['S6', 'Q_ALTER, Q_GESCHLECHT, Q_PLZ', 'Quota-Steuerung an den Anfang verschoben (vor Screening)', 'Intervista-Empfehlung'],
        ['AI-04', 'T7', 'Positive Formulierung: «Ich kann mich auf die Informationen von EnergieSchweiz verlassen.» statt invertierter Aussage', 'Meeting 2026-02-13'],
        ['AI-05', 'W4', '«organisieren und in die Wege zu leiten» statt «umzusetzen» — Self-Efficacy genauer erfassen', 'Meeting 2026-02-13'],
        ['AI-05', 'INTRO2', 'Kategorien überarbeitet', 'Meeting 2026-02-13'],
        ['AI-06', 'T0', '3-Stufen-Skala statt Ja/Nein: «Noch nie gehört / Schon gehört / Kenne ES»', 'Meeting 2026-02-13'],
        ['v3-Split', 'A1 → A1a, A1b', 'Getrennte Messung: A1a = Informationen, A1b = Werbung', 'EBF-Methodik'],
        ['v3-Split', 'A5 → A5a, A5b', 'Getrennte Messung: A5a = Gesellschaft, A5b = Umwelt', 'EBF-Methodik'],
        ['v3-Split', 'I1 → I1a, I1b', 'Getrennte Messung: I1a = Förderprogramme, I1b = Beratung', 'EBF-Methodik'],
        ['v3-NEU', 'BCM-Parameter', 'Neue Spalte: Verhaltensökonomische Parameter-Zuordnung pro Item', 'EBF-Framework'],
        ['v3-NEU', 'Literatur', 'Neue Spalte: Wissenschaftliche Referenzen pro Item', 'EBF-Framework'],
        ['v3-NEU', 'Hinweise', 'Neue Spalte: Änderungsnotizen und Begründungen', 'EBF-Framework'],
        ['v4-NEU', 'Z1', 'Von Rangfolge zu Likert-Matrix geändert: Wichtigkeit pro Faktor einzeln bewertbar', 'Messkonzept 2026'],
        ['v4-NEU', 'Z2', 'MaxDiff-Design für Utility-Gewichte (Best-Worst Scaling)', 'Louviere et al. (2015)'],
        ['v4-NEU', 'Modul-Tabs', 'Separate Tabellenblätter pro Modul (M1-M5) statt einem Gesamtblatt', 'Meeting 2026-02-18 (Andrea)'],
        ['v4-NEU', 'Kampagnen-Qs', 'Modul-spezifische Kampagnen-Fragen (Z5-Z9) für M3/M4/M5', 'BFE-Messkonzept'],
        ['v4-NEU', 'WILLKOMMEN', 'Modul-spezifischer Willkommenstext pro Tab', 'Modularisierung'],
        ['v4-STWE', '(geplant)', 'STWE-spezifische Fragen in späterem Update (Stockwerkeigentum)', 'Messkonzept 2026'],
    ]

    for i, change in enumerate(changes, 2):
        for col, val in enumerate(change, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER

    return ws


def create_tabelle1(wb):
    """Erstelle Tabelle1 mit Zielgruppen-Info."""
    ws = wb.create_sheet(title='Tabelle1')
    ws.cell(row=1, column=1, value='Zielgruppen-Steuerung').font = SECTION_FONT
    notes = [
        'Modul 1 (Gesamtmodernisierung): Eigentümer:innen EFH/MFH',
        'Modul 2 (Gebäudehülle): Eigentümer:innen EFH/MFH',
        'Modul 3 (Heizungsersatz): Eigentümer:innen EFH/MFH mit fossiler Heizung',
        'Modul 4 (Solarenergie): Eigentümer:innen EFH/MFH ohne PV',
        'Modul 5 (E-Mobilität): Eigentümer:innen EFH/MFH/STWE',
        'Zuweisung via INTRO-Screening (INTRO0-INTRO13, MODUL_ZUWEISUNG)',
    ]
    for i, note in enumerate(notes, 2):
        ws.cell(row=i, column=1, value=note).font = NORMAL_FONT
    ws.column_dimensions['A'].width = 70
    return ws


def create_moduluebersicht(wb):
    """Erstelle Modulübersicht-Tab."""
    ws = wb.create_sheet(title='Modulübersicht')
    headers = ['Modul', 'Thema', 'Zielgruppe', 'Basis-Items', 'Kampagnen-Qs', 'Total Items']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    modules_data = [
        ['M1', 'Gesamtmodernisierung', 'Eigentümer:innen EFH/MFH', '~25 (A+W+I+T+Z)', '0', '~25'],
        ['M2', 'Gebäudehülle', 'Eigentümer:innen EFH/MFH', '~25 (A+W+I+T+Z)', '0', '~25'],
        ['M3', 'Heizungsersatz', 'Eigentümer:innen mit fossiler Heizung', '~25 (A+W+I+T+Z)', '5 (Z5-Z9)', '~30'],
        ['M4', 'Solarenergie', 'Eigentümer:innen ohne PV', '~25 (A+W+I+T+Z)', '5 (Z5-Z9)', '~30'],
        ['M5', 'E-Mobilität', 'Eigentümer:innen EFH/MFH/STWE', '~25 (A+W+I+T+Z)', '3 (Z5-Z7)', '~28'],
    ]
    for i, row_data in enumerate(modules_data, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER

    for col_letter, width in {'A': 8, 'B': 25, 'C': 35, 'D': 22, 'E': 15, 'F': 14}.items():
        ws.column_dimensions[col_letter].width = width

    return ws


def main():
    """Hauptfunktion: Generiere BFE019 Fragebogen v4."""
    print("=" * 60)
    print("BFE019 Fragebogen v4 Generator")
    print("=" * 60)

    # V3 Template laden
    print(f"\n1. Lade v3 Template: {os.path.basename(V3_FILE)}")
    v3_wb = load_workbook(V3_FILE)
    print(f"   Sheets: {v3_wb.sheetnames}")

    # Neues Workbook erstellen
    wb = Workbook()

    # Tab 1: Modul 1 (komplett aus v3 kopiert)
    print("\n2. Kopiere Modul 1 aus v3 Template...")
    copy_v3_template(wb, v3_wb)
    print("   ✓ Modul 1 - Gesamtmodernisierung")

    # Tab 2-5: Module 2-5 generieren
    for mod_num in [2, 3, 4, 5]:
        config = MODULE_CONFIG[mod_num]
        print(f"\n3.{mod_num-1}. Generiere {config['tab_name']}...")
        create_module_tab(wb, mod_num, config)
        campaign_count = len(config.get('campaign_questions', []))
        print(f"   ✓ {config['name']} (Kampagnen-Fragen: {campaign_count})")

    # Tab 6: Modulübersicht
    print("\n4. Erstelle Modulübersicht...")
    create_moduluebersicht(wb)
    print("   ✓ Modulübersicht")

    # Tab 7: Tabelle1
    print("\n5. Erstelle Tabelle1 (Zielgruppen)...")
    create_tabelle1(wb)
    print("   ✓ Tabelle1")

    # Tab 8: Änderungsprotokoll
    print("\n6. Erstelle Änderungsprotokoll...")
    create_aenderungsprotokoll(wb)
    print("   ✓ Änderungsprotokoll (20 Einträge)")

    # Speichern
    print(f"\n7. Speichere: {os.path.basename(OUTPUT_FILE)}")
    wb.save(OUTPUT_FILE)
    print(f"   ✓ Gespeichert: {OUTPUT_FILE}")

    # Zusammenfassung
    print("\n" + "=" * 60)
    print("ZUSAMMENFASSUNG")
    print("=" * 60)
    print(f"Tabs: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name}: {ws.max_row} Zeilen × {ws.max_column} Spalten")
    print(f"\nDatei: {OUTPUT_FILE}")
    print(f"Grösse: wird nach Speicherung angezeigt")
    print("=" * 60)


if __name__ == '__main__':
    main()
