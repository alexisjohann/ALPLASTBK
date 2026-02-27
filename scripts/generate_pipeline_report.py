#!/usr/bin/env python3
"""
Generate Pipeline Reports (Excel + PDF)

Generates reports showing:
- Most recent leads (by created_at)
- Leads requiring immediate follow-up
- Priority tracking overview

Usage:
    python scripts/generate_pipeline_report.py
    python scripts/generate_pipeline_report.py --output-dir outputs/reports
    python scripts/generate_pipeline_report.py --format excel
    python scripts/generate_pipeline_report.py --format pdf
"""

import argparse
import yaml
from datetime import datetime, timedelta
from pathlib import Path
import subprocess
import sys

# =============================================================================
# CONFIGURATION
# =============================================================================

LEAD_DB_PATH = Path(__file__).parent.parent / "data" / "sales" / "lead-database.yaml"
DEFAULT_OUTPUT_DIR = Path(__file__).parent.parent / "outputs" / "reports"

# Priority weights
STAGE_PRIORITY = {
    "NEGOTIATION": 1,   # Highest priority
    "PROPOSAL": 2,
    "QUALIFIED": 3,
    "PROSPECT": 4,
    "ACTIVE": 5,
    "SUSPECT": 6,
    "DORMANT": 7,
    "REACTIVATION": 8,
    "WON": 9,
    "CHURNED": 10,
    "LOST": 10,
}

# =============================================================================
# DATA LOADING
# =============================================================================

def load_database():
    """Load the lead database YAML file."""
    with open(LEAD_DB_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_leads_for_followup(db):
    """Get leads that need follow-up, sorted by urgency."""
    today = datetime.now().date()
    leads = []

    for lead in db.get("leads", []):
        # Skip closed stages
        stage = lead.get("stage", "")
        if stage in ["CHURNED", "LOST", "WON"]:
            continue

        # Get next action date
        next_action = lead.get("next_action", {})
        action_date_str = next_action.get("date")

        if action_date_str:
            try:
                action_date = datetime.strptime(action_date_str, "%Y-%m-%d").date()
                days_until = (action_date - today).days
            except ValueError:
                days_until = 999
                action_date = None
        else:
            days_until = 999
            action_date = None

        # Calculate priority score (lower = more urgent)
        stage_score = STAGE_PRIORITY.get(stage, 10)

        # Opportunities in negotiation/proposal get boost
        opportunity = lead.get("opportunity") or lead.get("opportunities", [{}])[0] if lead.get("opportunities") else {}
        probability = opportunity.get("probability", 0) if opportunity else 0

        priority_score = (
            days_until * 10 +          # Time urgency
            stage_score * 5 +          # Stage importance
            (100 - probability)        # Deal probability (higher = lower score)
        )

        leads.append({
            "id": lead.get("id"),
            "company": lead.get("company", {}).get("name", "?"),
            "short_name": lead.get("company", {}).get("short_name", "?"),
            "stage": stage,
            "owner": lead.get("relationship", {}).get("owner", "?"),
            "next_action_date": action_date,
            "next_action_type": next_action.get("type", "-"),
            "next_action_desc": next_action.get("description", "-"),
            "days_until": days_until if days_until != 999 else None,
            "probability": probability,
            "priority_score": priority_score,
            "industry": lead.get("industry", "?"),
            "country": lead.get("headquarters", {}).get("country", "?"),
            "created_at": lead.get("created_at") or lead.get("created"),
            "fit_score": lead.get("fit_score"),
        })

    # Sort by priority score (lowest = most urgent)
    leads.sort(key=lambda x: x["priority_score"])
    return leads


def get_recent_leads(db, limit=10):
    """Get most recently created leads."""
    leads = db.get("leads", [])

    def get_timestamp(lead):
        return lead.get("created_at") or lead.get("created") or ""

    sorted_leads = sorted(leads, key=get_timestamp, reverse=True)

    result = []
    for lead in sorted_leads[:limit]:
        result.append({
            "id": lead.get("id"),
            "company": lead.get("company", {}).get("name", "?"),
            "short_name": lead.get("company", {}).get("short_name", "?"),
            "stage": lead.get("stage"),
            "owner": lead.get("relationship", {}).get("owner", "?"),
            "created_at": lead.get("created_at") or lead.get("created"),
            "industry": lead.get("industry", "?"),
            "country": lead.get("headquarters", {}).get("country", "?"),
        })

    return result


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_excel(db, output_path):
    """Generate Excel report with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ openpyxl nicht installiert. Installiere mit: pip install openpyxl")
        return False

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="024079", end_color="024079", fill_type="solid")
    urgent_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    warning_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    ok_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # -------------------------------------------------------------------------
    # Sheet 1: Follow-Up Priorität
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Follow-Up Priorität"

    headers1 = ["#", "Lead-ID", "Firma", "Stage", "Owner", "Nächste Aktion",
                "Datum", "Tage bis", "Typ", "Wahrsch.", "Branche", "Land"]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    followup_leads = get_leads_for_followup(db)

    for row, lead in enumerate(followup_leads, 2):
        ws1.cell(row=row, column=1, value=row-1).border = thin_border
        ws1.cell(row=row, column=2, value=lead["id"]).border = thin_border
        ws1.cell(row=row, column=3, value=lead["short_name"]).border = thin_border
        ws1.cell(row=row, column=4, value=lead["stage"]).border = thin_border
        ws1.cell(row=row, column=5, value=lead["owner"]).border = thin_border
        ws1.cell(row=row, column=6, value=lead["next_action_desc"][:50] if lead["next_action_desc"] else "-").border = thin_border

        date_cell = ws1.cell(row=row, column=7, value=lead["next_action_date"].strftime("%d.%m.%Y") if lead["next_action_date"] else "-")
        date_cell.border = thin_border

        days_cell = ws1.cell(row=row, column=8, value=lead["days_until"] if lead["days_until"] is not None else "-")
        days_cell.border = thin_border
        days_cell.alignment = Alignment(horizontal="center")

        # Color coding for urgency
        if lead["days_until"] is not None:
            if lead["days_until"] <= 0:
                days_cell.fill = urgent_fill
            elif lead["days_until"] <= 7:
                days_cell.fill = warning_fill
            elif lead["days_until"] <= 14:
                days_cell.fill = ok_fill

        ws1.cell(row=row, column=9, value=lead["next_action_type"]).border = thin_border

        prob_cell = ws1.cell(row=row, column=10, value=f"{lead['probability']}%" if lead["probability"] else "-")
        prob_cell.border = thin_border
        prob_cell.alignment = Alignment(horizontal="center")

        ws1.cell(row=row, column=11, value=lead["industry"]).border = thin_border
        ws1.cell(row=row, column=12, value=lead["country"]).border = thin_border

    # Adjust column widths
    col_widths1 = [4, 12, 25, 14, 8, 40, 12, 10, 12, 10, 18, 6]
    for i, width in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # -------------------------------------------------------------------------
    # Sheet 2: Neueste Leads
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet("Neueste Leads")

    headers2 = ["#", "Lead-ID", "Firma", "Stage", "Owner", "Erstellt", "Branche", "Land"]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    recent_leads = get_recent_leads(db, limit=20)

    for row, lead in enumerate(recent_leads, 2):
        ws2.cell(row=row, column=1, value=row-1).border = thin_border
        ws2.cell(row=row, column=2, value=lead["id"]).border = thin_border
        ws2.cell(row=row, column=3, value=lead["short_name"]).border = thin_border
        ws2.cell(row=row, column=4, value=lead["stage"]).border = thin_border
        ws2.cell(row=row, column=5, value=lead["owner"]).border = thin_border
        ws2.cell(row=row, column=6, value=lead["created_at"] or "-").border = thin_border
        ws2.cell(row=row, column=7, value=lead["industry"]).border = thin_border
        ws2.cell(row=row, column=8, value=lead["country"]).border = thin_border

    col_widths2 = [4, 12, 30, 14, 8, 24, 20, 6]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # -------------------------------------------------------------------------
    # Sheet 3: Pipeline Summary
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet("Pipeline Übersicht")

    summary = db.get("pipeline_summary", {})

    ws3.cell(row=1, column=1, value="Pipeline Report").font = Font(bold=True, size=14)
    ws3.cell(row=2, column=1, value=f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    ws3.cell(row=4, column=1, value="Stage").font = header_font
    ws3.cell(row=4, column=1).fill = header_fill
    ws3.cell(row=4, column=2, value="Anzahl").font = header_font
    ws3.cell(row=4, column=2).fill = header_fill

    by_stage = summary.get("by_stage", {})
    row = 5
    for stage, count in by_stage.items():
        ws3.cell(row=row, column=1, value=stage)
        ws3.cell(row=row, column=2, value=count)
        row += 1

    ws3.cell(row=row+1, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(row=row+1, column=2, value=summary.get("total_leads", 0)).font = Font(bold=True)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel erstellt: {output_path}")
    return True


# =============================================================================
# PDF GENERATION (via Markdown → PDF)
# =============================================================================

def generate_pdf(db, output_path):
    """Generate PDF report via Markdown conversion."""

    today = datetime.now()
    md_content = f"""# Sales Pipeline Report

**FehrAdvice & Partners AG**
Stand: {today.strftime('%d.%m.%Y %H:%M')}

---

## 🔥 Follow-Up Priorität (Top 20)

Die wichtigsten Leads zum sofortigen Nachverfolgen, sortiert nach Dringlichkeit.

| # | Lead | Firma | Stage | Owner | Nächste Aktion | Datum | Tage |
|---|------|-------|-------|-------|----------------|-------|------|
"""

    followup_leads = get_leads_for_followup(db)[:20]

    for i, lead in enumerate(followup_leads, 1):
        date_str = lead["next_action_date"].strftime("%d.%m.%Y") if lead["next_action_date"] else "-"
        days_str = str(lead["days_until"]) if lead["days_until"] is not None else "-"

        # Urgency indicator
        if lead["days_until"] is not None:
            if lead["days_until"] <= 0:
                urgency = "🔴"
            elif lead["days_until"] <= 7:
                urgency = "🟡"
            else:
                urgency = "🟢"
        else:
            urgency = "⚪"

        desc = (lead["next_action_desc"][:30] + "...") if lead["next_action_desc"] and len(lead["next_action_desc"]) > 30 else (lead["next_action_desc"] or "-")

        md_content += f"| {i} | {lead['id']} | {lead['short_name']} | {lead['stage']} | {lead['owner']} | {desc} | {date_str} | {urgency} {days_str} |\n"

    md_content += """
**Legende:** 🔴 Überfällig | 🟡 Diese Woche | 🟢 > 7 Tage | ⚪ Kein Datum

---

## 📅 Neueste Leads (Top 10)

Die zuletzt erstellten Leads in der Datenbank.

| # | Lead | Firma | Stage | Owner | Erstellt | Branche |
|---|------|-------|-------|-------|----------|---------|
"""

    recent_leads = get_recent_leads(db, limit=10)

    for i, lead in enumerate(recent_leads, 1):
        created = lead["created_at"] or "-"
        if "T" in str(created):
            # Parse ISO timestamp
            try:
                dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                created = dt.strftime("%d.%m.%Y %H:%M")
            except:
                pass

        md_content += f"| {i} | {lead['id']} | {lead['short_name']} | {lead['stage']} | {lead['owner']} | {created} | {lead['industry']} |\n"

    md_content += """

---

## 📊 Pipeline Übersicht

"""

    summary = db.get("pipeline_summary", {})
    by_stage = summary.get("by_stage", {})

    md_content += "| Stage | Anzahl |\n|-------|--------|\n"
    for stage, count in by_stage.items():
        md_content += f"| {stage} | {count} |\n"

    md_content += f"\n**Total: {summary.get('total_leads', 0)} Leads**\n"

    # BEATRIX Pipeline
    beatrix = summary.get("beatrix_pipeline", {})
    if beatrix:
        md_content += f"\n### BEATRIX Pipeline\n\n"
        md_content += f"Anzahl: {beatrix.get('count', 0)} Leads\n\n"
        md_content += f"Lead-IDs: {', '.join(beatrix.get('leads', []))}\n"

    md_content += """

---

*Automatisch generiert von FehrAdvice Sales Pipeline System*
"""

    # Write Markdown
    md_path = output_path.with_suffix(".md")
    md_path.parent.mkdir(parents=True, exist_ok=True)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"✅ Markdown erstellt: {md_path}")

    # Convert to PDF using pandoc (if available)
    try:
        result = subprocess.run(
            ["pandoc", str(md_path), "-o", str(output_path),
             "--pdf-engine=xelatex",
             "-V", "geometry:margin=2cm",
             "-V", "mainfont:DejaVu Sans"],
            capture_output=True,
            text=True
        )
        if result.returncode == 0:
            print(f"✅ PDF erstellt: {output_path}")
            return True
        else:
            print(f"⚠️ PDF-Konvertierung fehlgeschlagen: {result.stderr}")
            print(f"   Markdown-Datei verfügbar: {md_path}")
            return False
    except FileNotFoundError:
        print("⚠️ pandoc nicht installiert - nur Markdown erstellt")
        print(f"   Installiere pandoc für PDF: apt-get install pandoc texlive-xetex")
        return False


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Generate Pipeline Reports")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR,
                        help="Output directory for reports")
    parser.add_argument("--format", choices=["excel", "pdf", "both"], default="both",
                        help="Output format(s)")

    args = parser.parse_args()

    print("\n" + "="*60)
    print("  📊 PIPELINE REPORT GENERATOR")
    print("="*60 + "\n")

    # Load data
    db = load_database()
    total_leads = len(db.get("leads", []))
    print(f"📁 Geladen: {total_leads} Leads\n")

    # Generate timestamp for filenames
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")

    # Generate reports
    if args.format in ["excel", "both"]:
        excel_path = args.output_dir / f"pipeline_report_{timestamp}.xlsx"
        generate_excel(db, excel_path)

    if args.format in ["pdf", "both"]:
        pdf_path = args.output_dir / f"pipeline_report_{timestamp}.pdf"
        generate_pdf(db, pdf_path)

    print("\n" + "="*60)
    print("  ✅ FERTIG")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
