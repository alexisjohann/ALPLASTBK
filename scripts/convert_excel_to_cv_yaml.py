#!/usr/bin/env python3
"""Convert AWE002 Excel context vectors to YAML format.

Usage:
    python scripts/convert_excel_to_cv_yaml.py
"""

import os
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Column mappings per file (header -> yaml_key)
# We normalize headers across the 3 files
HEADER_MAP = {
    "ID": "id",
    "Dimension": "dimension",
    "Faktor": "faktor",
    "Definition": "definition",
    "Kontext": "kontext",
    "Kontext (AWE St. Gallen)": "kontext",
    "Kontext (Kanton SG)": "kontext",
    "Salienz": "salienz",
    "Salienz des Faktors": "salienz",
    "Operationalisierung / Indikator": "indikator",
    "3 Gründe": "gruende",
    "3 Gründe für den Datenpunkt": "gruende",
    "Datenpunkt": "datenpunkt",
    "Trend": "trend",
    "Unsicherheit": "unsicherheit",
    "Risiko / Chance": "risiko_chance",
    "Stärke / Schwäche": "staerke_schwaeche",
    "Beeinflussbarkeit": "beeinflussbarkeit",
    "Status": "status",
}


def yaml_escape(val):
    """Escape a value for safe YAML output."""
    if val is None:
        return "null"
    s = str(val).strip()
    if not s:
        return "null"
    # Numbers that should stay as-is
    try:
        float(s)
        return s
    except ValueError:
        pass
    # Quote strings that could cause YAML issues
    needs_quote = False
    if s.startswith(("{", "[", "'", '"', "&", "*", "!", "|", ">", "%", "@", "`")):
        needs_quote = True
    if ":" in s or "#" in s or "\n" in s:
        needs_quote = True
    if s.lower() in ("true", "false", "yes", "no", "null", "on", "off"):
        needs_quote = True
    if needs_quote or any(c in s for c in ['"']):
        # Use single quotes, escaping internal single quotes
        escaped = s.replace("'", "''")
        return f"'{escaped}'"
    # For long strings or strings with special chars, use quotes
    if len(s) > 80 or "," in s or ";" in s:
        escaped = s.replace("'", "''")
        return f"'{escaped}'"
    return s


def read_excel(filepath):
    """Read an Excel file and return header list and list of row dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Read headers
    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers.append((col, h.strip()))

    # Map headers to YAML keys
    col_map = {}
    for col_idx, header in headers:
        yaml_key = HEADER_MAP.get(header)
        if yaml_key:
            col_map[col_idx] = yaml_key
        else:
            print(f"  WARNING: Unknown header '{header}' in {filepath}")

    # Read data rows
    rows = []
    for row in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, yaml_key in col_map.items():
            val = ws.cell(row=row, column=col_idx).value
            row_data[yaml_key] = val
        # Skip empty rows
        if row_data.get("id"):
            rows.append(row_data)

    return rows


def group_by_dimension(rows):
    """Group rows by dimension, preserving order."""
    dims = []
    dim_map = {}
    for row in rows:
        d = row.get("dimension", "Unknown")
        if d not in dim_map:
            dim_map[d] = []
            dims.append(d)
        dim_map[d].append(row)
    return [(d, dim_map[d]) for d in dims]


def write_yaml(rows, metadata, output_path):
    """Write rows to a YAML file with metadata header."""
    grouped = group_by_dimension(rows)

    # Determine which keys are available (beyond id and dimension)
    all_keys = set()
    for row in rows:
        all_keys.update(k for k, v in row.items() if v is not None)

    # Ordered keys for output
    key_order = [
        "id", "dimension", "faktor", "definition", "kontext",
        "salienz", "indikator", "gruende", "datenpunkt",
        "trend", "unsicherheit", "risiko_chance", "staerke_schwaeche",
        "beeinflussbarkeit", "status"
    ]
    used_keys = [k for k in key_order if k in all_keys]

    lines = []
    lines.append("# -----------------------------------------------------------")
    lines.append(f"# {metadata['title']}")
    lines.append(f"# Projekt: {metadata['projekt']}")
    lines.append(f"# Ebene: {metadata['ebene']}")
    lines.append(f"# Quelle: {metadata['quelle']}")
    lines.append(f"# Konvertiert: {date.today().isoformat()}")
    lines.append(f"# Faktoren: {len(rows)}")
    lines.append(f"# Dimensionen: {len(grouped)}")
    lines.append("# -----------------------------------------------------------")
    lines.append("")
    lines.append("metadata:")
    lines.append(f"  titel: {yaml_escape(metadata['title'])}")
    lines.append(f"  projekt: {yaml_escape(metadata['projekt'])}")
    lines.append(f"  ebene: {yaml_escape(metadata['ebene'])}")
    lines.append(f"  quelle: {yaml_escape(metadata['quelle'])}")
    lines.append(f"  konvertiert: '{date.today().isoformat()}'")
    lines.append(f"  faktoren_total: {len(rows)}")
    lines.append(f"  dimensionen_total: {len(grouped)}")
    lines.append("  dimensionen:")
    for dim_name, dim_rows in grouped:
        lines.append(f"    - name: {yaml_escape(dim_name)}")
        lines.append(f"      faktoren: {len(dim_rows)}")
    lines.append("")
    lines.append("faktoren:")

    for dim_name, dim_rows in grouped:
        lines.append("")
        lines.append(f"  # === {dim_name} ({len(dim_rows)} Faktoren) ===")
        for row in dim_rows:
            lines.append("")
            lines.append(f"  - id: {yaml_escape(row.get('id'))}")
            lines.append(f"    dimension: {yaml_escape(row.get('dimension'))}")
            for key in used_keys:
                if key in ("id", "dimension"):
                    continue
                val = row.get(key)
                lines.append(f"    {key}: {yaml_escape(val)}")

    lines.append("")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  Written: {output_path} ({len(rows)} factors, {len(grouped)} dimensions)")


def main():
    base = "data/customers/awe-sg/projects/AWE002"
    out_dir = os.path.join(base, "Kontextvektoren")

    conversions = [
        {
            "input": os.path.join(base, "AWE001_Projektvektor_2026-02-25.xlsx"),
            "output": os.path.join(out_dir, "CV_AWE002_00_projektvektor.yaml"),
            "metadata": {
                "title": "AWE002 Projektvektor — Energiekonzept St. Gallen",
                "projekt": "AWE002",
                "ebene": "Projekt",
                "quelle": "AWE001_Projektvektor_2026-02-25.xlsx",
            },
        },
        {
            "input": os.path.join(base, "AWE_Unternehemen_Kontextvektor.xlsx"),
            "output": os.path.join(out_dir, "CV_AWE002_01_unternehmen.yaml"),
            "metadata": {
                "title": "AWE Unternehmens-Kontextvektor — Amt für Wasser und Energie SG",
                "projekt": "AWE002",
                "ebene": "Organisation",
                "quelle": "AWE_Unternehemen_Kontextvektor.xlsx",
            },
        },
        {
            "input": os.path.join(base, "SG_Kanton_Kontextvektor.xlsx"),
            "output": os.path.join(out_dir, "CV_AWE002_02_kanton_sg.yaml"),
            "metadata": {
                "title": "Kantons-Kontextvektor St. Gallen — Energiepolitischer Rahmen",
                "projekt": "AWE002",
                "ebene": "Kanton",
                "quelle": "SG_Kanton_Kontextvektor.xlsx",
            },
        },
    ]

    total_factors = 0
    for conv in conversions:
        print(f"\nConverting: {conv['input']}")
        rows = read_excel(conv["input"])
        write_yaml(rows, conv["metadata"], conv["output"])
        total_factors += len(rows)

    print(f"\nTotal: {total_factors} factors converted across {len(conversions)} files.")


if __name__ == "__main__":
    main()
